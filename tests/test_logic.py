import unittest
from datetime import datetime
import os
import tempfile

import openpyxl

from src.logic import OrderParser, CatalogParser, PriceMatcher
from src.writer import StatementWriter
from src.utils import extract_quantity, normalize_string, parse_date_from_sheet_name, quantities_equal

class TestLogic(unittest.TestCase):
    def test_normalize(self):
        self.assertEqual(normalize_string("  Apple  "), "apple")
        self.assertEqual(normalize_string("Apple Pie"), "applepie")

    def test_matcher_exact(self):
        catalog = {
            "apple": [{"spec": "1kg", "price": 1000}, {"spec": "2kg", "price": 1800}]
        }
        matcher = PriceMatcher(catalog)
        
        price = matcher.get_price("Apple", "1kg")
        self.assertEqual(price, 1000)
        
        price = matcher.get_price("Apple", "2kg")
        self.assertEqual(price, 1800)
        
        price = matcher.get_price("Apple", "500g")
        self.assertIsNone(price)

    def test_matcher_partial_name(self):
        # "Name exact match first, if absent, try partial"
        catalog = {
            "GalaApple": [{"spec": "1kg", "price": 1200}]
        }
        matcher = PriceMatcher(catalog)
        
        # Exact match fails for "Apple", but "Apple" in "GalaApple" -> Partial?
        # My implementation: if search_term in catalog_key OR catalog_key in search_term
        
        # Test 1: Search "Apple", Catalog "GalaApple" -> "Apple" in "GalaApple" -> NO Match (Strict)
        price = matcher.get_price("Apple", "1kg")
        self.assertIsNone(price)

    def test_matcher_strategies(self):
        catalog = {
            "pork": [
                {"spec": "1kg,korean", "price": 5000},
                {"spec": "500g", "price": 3000}
            ],
            "onion": [
                {"spec": "10kg", "price": 2000}
            ]
        }
        matcher = PriceMatcher(catalog)
        
        # Strategy 3: Token/Permutation "korean, 1kg" vs "1kg,korean"
        price = matcher.get_price("pork", "korean, 1kg")
        self.assertEqual(price, 5000)
        
        # Strategy 4: Blank Spec "onion (None)" -> "onion (10kg)" (Unique)
        price = matcher.get_price("onion", None)
        self.assertEqual(price, 2000)
        
        # Strategy 4: Blank Spec "pork (None)" -> Multiple choices -> None
        price = matcher.get_price("pork", "")
        self.assertIsNone(price)
        
        # Strategy 5: Partial Spec "pork (1kg)" -> "pork (1kg,korean)"
        price = matcher.get_price("pork", "1kg")
        self.assertEqual(price, 5000)

    def test_quantity_equivalence(self):
        self.assertTrue(quantities_equal(extract_quantity("1kg"), extract_quantity("1000 g")))
        self.assertTrue(quantities_equal(extract_quantity("500g x 2"), extract_quantity("1kg")))
        self.assertTrue(quantities_equal(extract_quantity("1L"), extract_quantity("1000ml")))

    def test_date_parser_uses_fallback_year(self):
        self.assertEqual(parse_date_from_sheet_name("4.28", fallback_year=2027), datetime(2027, 4, 28))
        self.assertEqual(parse_date_from_sheet_name("2028년 5월 3일"), datetime(2028, 5, 3))

    def test_catalog_parser_accepts_header_aliases(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "catalog.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["B3"] = "제품명"
            ws["C3"] = "용량"
            ws["D3"] = "단가(원)"
            ws["B4"] = "국내산 양파"
            ws["C4"] = "1kg"
            ws["D4"] = "1,200원"
            wb.save(path)

            price_map = CatalogParser(path).parse()
            matcher = PriceMatcher(price_map)

            self.assertEqual(matcher.get_price("양파", "1000g"), 1200)

    def test_order_parser_detects_shifted_headers(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "order.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "4.28"
            ws["C4"] = "일반 발주"
            ws["C5"] = "제품명"
            ws["D5"] = "용량"
            ws["E5"] = "단위"
            ws["F5"] = "수량"
            ws["C6"] = "양파"
            ws["D6"] = "1kg"
            ws["E6"] = "봉"
            ws["F6"] = 3
            wb.save(path)

            parsed = OrderParser(path).parse_sheet("4.28")

            self.assertEqual(len(parsed["default"]), 1)
            self.assertEqual(parsed["default"][0]["name"], "양파")
            self.assertEqual(parsed["default"][0]["qty"], 3)

    def test_writer_keeps_template_header(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            writer = StatementWriter(temp_dir)
            out_path = writer.write_statement(
                [{
                    "no": 1,
                    "name": "양파",
                    "spec": "1kg",
                    "unit": "봉",
                    "qty": 2,
                    "price": 1200,
                    "match": {"status": "matched"},
                }],
                "default",
                datetime(2027, 4, 28),
            )

            wb = openpyxl.load_workbook(out_path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=9, column=2).value, "품목명")
            self.assertEqual(ws.cell(row=10, column=2).value, "양파")

if __name__ == '__main__':
    unittest.main()
