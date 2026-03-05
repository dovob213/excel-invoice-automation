import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def create_template(filename, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Set up headers
    headers = ["NO", "품목명", "규격", "단위", "수량", "단가", "금액", "비고"]
    
    # Title
    ws['A1'] = title
    ws['A1'].font = Font(size=20, bold=True)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Info rows (simplified for dummy)
    ws['A4'] = "수신: "
    ws['F4'] = "발신: "
    ws['A5'] = "참조: "
    ws['F5'] = "날짜: "
    
    # Header row at 9
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='thin'))

    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    create_template("templates/transaction_statement.xlsx", "거래명세서")
    create_template("templates/transaction_statement_employee.xlsx", "거래명세서(직원용)")
    create_template("templates/transaction_statement_dairy.xlsx", "거래명세서(유제품)")
