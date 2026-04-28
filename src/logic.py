from dataclasses import dataclass
from difflib import SequenceMatcher
import re

import openpyxl

from src.utils import (
    compact_name,
    extract_quantity,
    header_field,
    infer_category,
    normalize_header,
    normalize_name,
    normalize_spec,
    parse_number,
    quantities_equal,
    spec_tokens,
    tokenize_name,
)


SECTION_ORDER = ["default", "employee", "dairy"]
SECTION_LABELS = {
    "default": "일반",
    "employee": "직원용",
    "dairy": "유제품",
}
AUTO_MATCH_THRESHOLD = 86
REVIEW_MATCH_THRESHOLD = 72


@dataclass
class TableBlock:
    section: str
    header_row: int
    columns: dict
    start_col: int
    end_col: int


class OrderParser:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def parse_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        blocks = self._find_table_blocks(sheet)

        parsed_data = {
            "default": [],
            "employee": [],
            "dairy": [],
        }

        for block in blocks:
            items = self._parse_block(sheet, block)
            parsed_data.setdefault(block.section, []).extend(items)

        return parsed_data

    def _find_table_blocks(self, sheet):
        blocks = []
        seen = set()

        for row_idx in range(1, sheet.max_row + 1):
            header_cells = []
            for col_idx in range(1, sheet.max_column + 1):
                field = header_field(sheet.cell(row=row_idx, column=col_idx).value)
                if field in {"no", "name", "spec", "unit", "qty", "category"}:
                    header_cells.append((col_idx, field))

            if not any(field == "name" for _, field in header_cells):
                continue
            if not any(field in {"qty", "spec", "unit", "no"} for _, field in header_cells):
                continue

            groups = self._split_header_groups(header_cells)
            valid_groups = []
            for group in groups:
                columns = self._columns_from_group(group)
                if "name" not in columns:
                    continue
                if not any(field in columns for field in ("qty", "spec", "unit", "no")):
                    continue
                valid_groups.append((group, columns))

            for order, (group, columns) in enumerate(valid_groups):
                start_col = min(col for col, _ in group)
                end_col = max(col for col, _ in group)
                key = (row_idx, tuple(sorted(columns.items())))
                if key in seen:
                    continue
                seen.add(key)

                section = self._detect_section(sheet, row_idx, start_col, end_col, order)
                blocks.append(TableBlock(section, row_idx, columns, start_col, end_col))

        if blocks:
            return sorted(blocks, key=lambda block: (block.header_row, block.start_col))

        return self._fallback_blocks(sheet)

    def _split_header_groups(self, header_cells):
        groups = []
        current = []
        current_fields = set()
        previous_col = None

        for col_idx, field in header_cells:
            should_split = False
            if current and previous_col is not None:
                should_split = col_idx - previous_col > 3
            if current and field in current_fields and field in {"no", "name", "qty"}:
                should_split = True

            if should_split:
                groups.append(current)
                current = []
                current_fields = set()

            current.append((col_idx, field))
            current_fields.add(field)
            previous_col = col_idx

        if current:
            groups.append(current)
        return groups

    def _columns_from_group(self, group):
        columns = {}
        for col_idx, field in group:
            columns.setdefault(field, col_idx)
        return columns

    def _detect_section(self, sheet, header_row, start_col, end_col, fallback_index):
        text_parts = []
        for row_idx in range(max(1, header_row - 3), header_row + 1):
            for col_idx in range(max(1, start_col - 2), min(sheet.max_column, end_col + 2) + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value:
                    text_parts.append(str(value))

        context = normalize_name(" ".join(text_parts))
        if "직원" in context or "employee" in context:
            return "employee"
        if "유제품" in context or "우유" in context or "dairy" in context:
            return "dairy"
        if fallback_index < len(SECTION_ORDER):
            return SECTION_ORDER[fallback_index]
        return "default"

    def _fallback_blocks(self, sheet):
        max_col = sheet.max_column
        candidates = [
            TableBlock("default", 5, {"no": 1, "name": 2, "spec": 3, "unit": 4, "qty": 5}, 1, 7),
            TableBlock("employee", 5, {"no": 13, "name": 14, "spec": 15, "unit": 16, "qty": 17}, 13, 24),
            TableBlock("dairy", 5, {"no": 25, "name": 26, "spec": 27, "unit": 28, "qty": 29}, 25, 36),
        ]
        return [block for block in candidates if block.start_col <= max_col]

    def _parse_block(self, sheet, block):
        items = []
        empty_streak = 0
        max_empty_streak = 8

        for row_idx in range(block.header_row + 1, sheet.max_row + 1):
            row_values = {
                field: sheet.cell(row=row_idx, column=col_idx).value
                for field, col_idx in block.columns.items()
            }

            meaningful_values = [
                row_values.get("no"),
                row_values.get("name"),
                row_values.get("spec"),
                row_values.get("unit"),
                row_values.get("qty"),
            ]
            if all(value in (None, "") for value in meaningful_values):
                empty_streak += 1
                if empty_streak >= max_empty_streak:
                    break
                continue

            empty_streak = 0
            name = row_values.get("name")
            if not name:
                continue

            normalized_name = normalize_header(name)
            if normalized_name in {"식품명", "품목명", "제품명", "상품명", "품명"}:
                continue
            if any(word in normalize_name(name) for word in ("합계", "총계", "소계")):
                continue

            spec = row_values.get("spec")
            unit = row_values.get("unit")
            qty = row_values.get("qty")
            category = infer_category(
                name=name,
                spec=spec,
                explicit=row_values.get("category"),
                section=SECTION_LABELS.get(block.section, block.section),
            )

            items.append({
                "no": row_values.get("no") or len(items) + 1,
                "name": name,
                "spec": spec,
                "unit": unit,
                "qty": qty,
                "category": category,
                "section": block.section,
                "source_sheet": sheet.title,
                "source_row": row_idx,
            })

        return items


class CatalogParser:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)
        self.price_map = {}
        self.items = []

    def parse(self):
        self.price_map = {}
        self.items = []

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            blocks = self._find_catalog_blocks(sheet)
            for block in blocks:
                self._parse_catalog_block(sheet, block)

        return self.price_map

    def _find_catalog_blocks(self, sheet):
        blocks = []
        seen = set()

        for row_idx in range(1, min(sheet.max_row, 80) + 1):
            header_cells = []
            for col_idx in range(1, sheet.max_column + 1):
                field = header_field(sheet.cell(row=row_idx, column=col_idx).value)
                if field in {"name", "spec", "unit", "price", "category"}:
                    header_cells.append((col_idx, field))

            if not any(field == "name" for _, field in header_cells):
                continue
            if not any(field == "price" for _, field in header_cells):
                continue

            for group in self._split_header_groups(header_cells):
                columns = self._columns_from_group(group)
                if "name" not in columns or "price" not in columns:
                    continue
                key = (row_idx, tuple(sorted(columns.items())))
                if key in seen:
                    continue
                seen.add(key)
                start_col = min(col for col, _ in group)
                end_col = max(col for col, _ in group)
                blocks.append(TableBlock("catalog", row_idx, columns, start_col, end_col))

        return sorted(blocks, key=lambda block: (block.header_row, block.start_col))

    def _split_header_groups(self, header_cells):
        groups = []
        current = []
        current_fields = set()

        for col_idx, field in header_cells:
            if current and field in current_fields and field in {"name", "price"}:
                groups.append(current)
                current = []
                current_fields = set()
            current.append((col_idx, field))
            current_fields.add(field)

        if current:
            groups.append(current)
        return groups

    def _columns_from_group(self, group):
        return OrderParser._columns_from_group(self, group)

    def _parse_catalog_block(self, sheet, block):
        empty_streak = 0

        for row_idx in range(block.header_row + 1, sheet.max_row + 1):
            values = {
                field: sheet.cell(row=row_idx, column=col_idx).value
                for field, col_idx in block.columns.items()
            }
            name_val = values.get("name")
            price_val = parse_number(values.get("price"))

            if all(values.get(field) in (None, "") for field in ("name", "spec", "unit", "price")):
                empty_streak += 1
                if empty_streak >= 10:
                    break
                continue

            empty_streak = 0
            if not name_val or price_val is None:
                continue
            if any(word in normalize_name(name_val) for word in ("합계", "총계", "소계")):
                continue

            spec_val = values.get("spec") or values.get("unit") or ""
            category = infer_category(
                name=name_val,
                spec=spec_val,
                explicit=values.get("category"),
                section=sheet.title,
            )
            item = self._build_catalog_item(name_val, spec_val, price_val, category, sheet.title, row_idx)
            self.items.append(item)
            self.price_map.setdefault(item["name_key"], []).append(item)

    def _build_catalog_item(self, name, spec, price, category, sheet_name, row_idx):
        return {
            "name_key": compact_name(name),
            "name_normalized": normalize_name(name),
            "name_tokens": tokenize_name(name),
            "spec": normalize_spec(spec),
            "spec_tokens": spec_tokens(spec),
            "quantity": extract_quantity(spec),
            "price": price,
            "category": category,
            "original_spec": spec,
            "original_name": name,
            "sheet": sheet_name,
            "source_row": row_idx,
        }


class PriceMatcher:
    def __init__(self, price_catalogs):
        self.catalogs = price_catalogs
        self.items = self._flatten_catalogs(price_catalogs)

    def _flatten_catalogs(self, price_catalogs):
        items = []
        if isinstance(price_catalogs, list):
            source_items = price_catalogs
        else:
            source_items = []
            for name_key, candidates in price_catalogs.items():
                for candidate in candidates:
                    item = dict(candidate)
                    item.setdefault("name_key", name_key)
                    item.setdefault("original_name", name_key)
                    source_items.append(item)

        for candidate in source_items:
            item = dict(candidate)
            original_name = item.get("original_name") or item.get("name") or item.get("name_key") or ""
            original_spec = item.get("original_spec") or item.get("spec") or ""
            item["name_key"] = compact_name(original_name)
            item["name_normalized"] = normalize_name(original_name)
            item["name_tokens"] = tokenize_name(original_name)
            item["spec"] = normalize_spec(original_spec)
            item["spec_tokens"] = spec_tokens(original_spec)
            item["quantity"] = extract_quantity(original_spec)
            item.setdefault("category", infer_category(original_name, original_spec))
            item.setdefault("original_name", original_name)
            item.setdefault("original_spec", original_spec)
            items.append(item)
        return items

    def get_price(self, name, spec):
        result = self.match(name, spec)
        return result["price"] if result["status"] == "matched" else None

    def match(self, name, spec, category=None):
        order_name_key = compact_name(name)
        order_tokens = tokenize_name(name)
        order_spec = normalize_spec(spec)
        order_spec_tokens = spec_tokens(spec)
        order_quantity = extract_quantity(spec)
        order_category = infer_category(name=name, spec=spec, explicit=category)

        if not order_name_key:
            return self._empty_result("unmatched", "제품명이 비어 있습니다.")

        scored = []
        for item in self.items:
            name_score, name_reason = self._score_name(order_name_key, order_tokens, item)
            if name_score < 45:
                continue

            spec_score, spec_reason = self._score_spec(
                order_spec,
                order_spec_tokens,
                order_quantity,
                item,
            )
            category_score, category_reason = self._score_category(order_category, item.get("category"))
            confidence = round(name_score * 0.64 + spec_score * 0.31 + category_score * 0.05, 1)

            scored.append({
                "item": item,
                "confidence": confidence,
                "name_score": name_score,
                "spec_score": spec_score,
                "category_score": category_score,
                "reason": f"{name_reason}; {spec_reason}; {category_reason}",
            })

        if not scored:
            return self._empty_result("unmatched", "카탈로그에서 유사한 제품명을 찾지 못했습니다.")

        scored.sort(key=lambda entry: entry["confidence"], reverse=True)
        best = scored[0]
        second = scored[1] if len(scored) > 1 else None
        ambiguous = bool(second and best["confidence"] < 95 and best["confidence"] - second["confidence"] < 3)

        if best["confidence"] >= AUTO_MATCH_THRESHOLD and not ambiguous:
            status = "matched"
            price = best["item"].get("price")
            reason = best["reason"]
        elif best["confidence"] >= REVIEW_MATCH_THRESHOLD:
            status = "review"
            price = None
            reason = "검토 필요: 후보가 충분히 비슷하지만 자동 확정 기준에는 부족합니다. " + best["reason"]
        else:
            status = "unmatched"
            price = None
            reason = "자동 매칭 기준 미달입니다. " + best["reason"]

        if ambiguous and best["confidence"] >= REVIEW_MATCH_THRESHOLD:
            status = "review"
            price = None
            reason = "검토 필요: 비슷한 후보가 여러 개 있습니다. " + best["reason"]

        return {
            "status": status,
            "price": price,
            "suggested_price": best["item"].get("price"),
            "confidence": best["confidence"],
            "match_type": status,
            "reason": reason,
            "catalog_name": best["item"].get("original_name"),
            "catalog_spec": best["item"].get("original_spec"),
            "catalog_category": best["item"].get("category"),
            "candidates": [self._candidate_summary(entry) for entry in scored[:3]],
        }

    def _empty_result(self, status, reason):
        return {
            "status": status,
            "price": None,
            "suggested_price": None,
            "confidence": 0,
            "match_type": status,
            "reason": reason,
            "catalog_name": None,
            "catalog_spec": None,
            "catalog_category": None,
            "candidates": [],
        }

    def _candidate_summary(self, entry):
        item = entry["item"]
        return {
            "name": item.get("original_name"),
            "spec": item.get("original_spec"),
            "price": item.get("price"),
            "category": item.get("category"),
            "confidence": entry["confidence"],
        }

    def _score_name(self, order_name_key, order_tokens, item):
        catalog_key = item.get("name_key") or ""
        if order_name_key == catalog_key:
            return 100, "제품명 정확히 일치"

        if order_name_key and catalog_key and (order_name_key in catalog_key or catalog_key in order_name_key):
            shorter = order_name_key if len(order_name_key) <= len(catalog_key) else catalog_key
            if re.search(r"[가-힣]", shorter) and len(shorter) >= 2:
                return 84, "제품명 포함 관계"
            return 68, "제품명 일부 포함"

        catalog_tokens = item.get("name_tokens") or set()
        if order_tokens and catalog_tokens:
            overlap = len(order_tokens & catalog_tokens)
            if overlap:
                ratio = overlap / max(len(order_tokens), len(catalog_tokens))
                return 68 + ratio * 24, "제품명 토큰 일부 일치"

        ratio = SequenceMatcher(None, order_name_key, catalog_key).ratio()
        return ratio * 82, "제품명 유사도 비교"

    def _score_spec(self, order_spec, order_tokens, order_quantity, item):
        catalog_spec = item.get("spec") or ""
        catalog_quantity = item.get("quantity")

        if not order_spec and not catalog_spec:
            return 85, "규격 정보 없음"
        if order_spec and catalog_spec and order_spec == catalog_spec:
            return 100, "규격 정확히 일치"
        if quantities_equal(order_quantity, catalog_quantity):
            return 96, "용량 환산 일치"
        if order_quantity and catalog_quantity and order_quantity.get("unit") == catalog_quantity.get("unit"):
            return 5, "용량 불일치"
        if not order_spec:
            return 70, "발주서 규격 없음"
        if not catalog_spec:
            return 58, "카탈로그 규격 없음"

        catalog_tokens = item.get("spec_tokens") or set()
        if order_tokens and catalog_tokens and order_tokens == catalog_tokens:
            return 94, "규격 토큰 순서만 다름"
        if order_spec in catalog_spec or catalog_spec in order_spec:
            return 82, "규격 포함 관계"
        if order_tokens and catalog_tokens:
            overlap = len(order_tokens & catalog_tokens)
            if overlap:
                ratio = overlap / max(len(order_tokens), len(catalog_tokens))
                return 55 + ratio * 25, "규격 토큰 일부 일치"

        return 35, "규격 유사도 낮음"

    def _score_category(self, order_category, catalog_category):
        if not order_category or order_category == "기타" or not catalog_category:
            return 60, "카테고리 참고 정보 부족"
        if order_category == catalog_category:
            return 100, "카테고리 일치"
        return 45, "카테고리 다름"
