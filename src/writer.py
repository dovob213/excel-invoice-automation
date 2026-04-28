import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime, timedelta
import os

from src.utils import get_resource_path
from src.utils import parse_number

class StatementWriter:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
    def _load_template(self, template_type):
        # Determine template file based on type
        # template_type: 'default', 'employee', 'dairy'
        filename_map = {
            'default': 'transaction_statement.xlsx',
            'employee': 'transaction_statement_employee.xlsx',
            'dairy': 'transaction_statement_dairy.xlsx'
        }
        
        # Use resource path handler
        template_path = get_resource_path(os.path.join('templates', filename_map.get(template_type, 'transaction_statement.xlsx')))
        
        if not os.path.exists(template_path):
            # Fallback: check if it's in the current directory (for some Windows cases)
            fallback = os.path.join(os.getcwd(), 'templates', filename_map.get(template_type, 'transaction_statement.xlsx'))
            if os.path.exists(fallback):
                template_path = fallback
            else:
                raise FileNotFoundError(f"Template not found: {template_path}")
            
        return openpyxl.load_workbook(template_path)

    def _find_header_row(self, ws):
        for row in range(1, min(ws.max_row, 30) + 1):
            values = [ws.cell(row=row, column=col).value for col in range(1, min(ws.max_column, 12) + 1)]
            normalized = {str(value).replace(" ", "") for value in values if value is not None}
            if "품목명" in normalized and "단가" in normalized:
                return row
        return 9

    def _unique_output_path(self, path):
        if not os.path.exists(path):
            return path
        root, ext = os.path.splitext(path)
        counter = 1
        while True:
            candidate = f"{root}_{counter}{ext}"
            if not os.path.exists(candidate):
                return candidate
            counter += 1

    def write_statement(self, data_list, template_type, date_obj):
        """
        data_list: list of dicts {'no', 'name', 'spec', 'unit', 'qty', 'price', ...}
        template_type: 'default', 'employee', 'dairy'
        date_obj: datetime object for the "Receiving Date" (Input Date)
        """
        wb = self._load_template(template_type)
        ws = wb.active
        
        # Write Dates
        # Valid locations based on dummy template: F5 (Date)
        # However, requirements say: "Input Date (selected)" and "Sending Date (Input - 1)"
        # Let's put Input Date in "입고일" and Sending Date in "발송일" if slots exist.
        # User said: "Header: Sender/Receiver/Contact/SendingDate/ReceivingDate"
        # My dummy template has A4..F5.
        # Let's assume standard locations or just fill "Date" with Input Date for now.
        # User said: "Sending date = Input - 1".
        
        base_date = date_obj or datetime.now()
        sending_date = base_date - timedelta(days=1)
        date_str = base_date.strftime("%Y-%m-%d")
        sending_str = sending_date.strftime("%Y-%m-%d")
        
        # In dummy template, F5 is "Date: ". Let's append there? 
        # Or just find the cell and set it.
        # "발송일", "입고일" exact location might need config unless template is fixed.
        # I'll just put date in F5 for now as a placeholder.
        ws['F5'] = f"날짜: {date_str} (발송: {sending_str})"
        
        header_row = self._find_header_row(ws)
        start_row = header_row + 1
        
        # Style definition for missing price
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        for i, item in enumerate(data_list):
            current_row = start_row + i
            
            # Map item keys to columns (1-indexed)
            # 1: No, 2: Name, 3: Spec, 4: Unit, 5: Qty, 6: Price, 7: Amount, 8: Note
            
            ws.cell(row=current_row, column=1, value=item.get('no'))
            ws.cell(row=current_row, column=2, value=item.get('name'))
            ws.cell(row=current_row, column=3, value=item.get('spec'))
            ws.cell(row=current_row, column=4, value=item.get('unit'))
            ws.cell(row=current_row, column=5, value=item.get('qty'))
            
            price = item.get('price')
            qty = parse_number(item.get('qty'))
            match = item.get('match') or {}
            
            # Handle Price
            if price is not None:
                ws.cell(row=current_row, column=6, value=price)
                # Calculate Amount
                try:
                    amount = float(price) * float(qty) if qty else 0
                    ws.cell(row=current_row, column=7, value=amount)
                except (ValueError, TypeError):
                    ws.cell(row=current_row, column=7, value=0)
            else:
                # Missing Price -> Yellow Highlight + Empty
                ws.cell(row=current_row, column=6).fill = yellow_fill
                ws.cell(row=current_row, column=6).value = None # Leave blank
                ws.cell(row=current_row, column=7).value = None # Leave blank

            note = item.get('note')
            if not note and match:
                if match.get('status') == 'review':
                    note = f"검토: {match.get('catalog_name') or ''} {match.get('catalog_spec') or ''} / 제안단가 {match.get('suggested_price') or ''} / {match.get('confidence')}점"
                elif match.get('status') == 'unmatched':
                    note = f"미매칭: {match.get('reason') or ''}"
            if note:
                ws.cell(row=current_row, column=8, value=note)
                if match.get('status') == 'review':
                    ws.cell(row=current_row, column=8).fill = orange_fill
            
            # Apply styling to all cells in row
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                # Alignment?
                if col in [1, 4, 5, 6, 7]: # Numbers centric
                    cell.alignment = Alignment(horizontal='center')
        
        # Generate Output Filename
        # Format: "거래명세서_[Type]_[Date].xlsx"
        name_map = {
            'employee': '_직원용',
            'dairy': '_유제품'
        }
        type_suffix = name_map.get(template_type, "")
        filename = f"거래명세서{type_suffix}_{date_str}.xlsx"
        output_path = self._unique_output_path(os.path.join(self.output_dir, filename))
        
        wb.save(output_path)
        return output_path

    def write_review_report(self, data_list, date_obj):
        base_date = date_obj or datetime.now()
        date_str = base_date.strftime("%Y-%m-%d")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "매칭검토"

        headers = [
            "섹션", "상태", "신뢰도", "발주서 품목명", "발주서 규격", "수량", "카테고리",
            "확정 단가", "제안 단가", "카탈로그 품목명", "카탈로그 규격", "사유", "후보",
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        status_fills = {
            "matched": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
            "review": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "unmatched": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        }

        for item in data_list:
            match = item.get("match") or {}
            candidates = []
            for candidate in match.get("candidates", []):
                candidates.append(
                    f"{candidate.get('name') or ''} / {candidate.get('spec') or ''} / "
                    f"{candidate.get('price') or ''} / {candidate.get('confidence') or ''}점"
                )

            row = [
                item.get("section"),
                match.get("status"),
                match.get("confidence"),
                item.get("name"),
                item.get("spec"),
                item.get("qty"),
                item.get("category"),
                item.get("price"),
                match.get("suggested_price"),
                match.get("catalog_name"),
                match.get("catalog_spec"),
                match.get("reason"),
                "\n".join(candidates),
            ]
            ws.append(row)
            fill = status_fills.get(match.get("status"))
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill

        widths = [12, 12, 10, 24, 18, 10, 14, 12, 12, 24, 18, 48, 56]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        filename = f"매칭_검토리포트_{date_str}.xlsx"
        output_path = self._unique_output_path(os.path.join(self.output_dir, filename))
        wb.save(output_path)
        return output_path
